"""
Streaming Excel loader using openpyxl in read_only mode.
Memory-efficient for large sheets (500k+ rows).
"""
import openpyxl
from pathlib import Path
from typing import Iterator, Dict, Any, List, Optional
import structlog

logger = structlog.get_logger()


class StreamingExcelLoader:
    """Load Excel sheets in streaming mode (read_only, row-by-row)."""
    
    def __init__(self, file_path: str):
        """
        Initialize loader.
        
        Args:
            file_path: Path to Excel file
        """
        self.file_path = Path(file_path)
        if not self.file_path.exists():
            raise FileNotFoundError(f"Excel file not found: {file_path}")
        self.workbook = None
    
    def __enter__(self):
        """Context manager entry."""
        self.workbook = openpyxl.load_workbook(
            self.file_path,
            read_only=True,
            data_only=False  # Keep formulas/formats for inspection
        )
        return self
    
    def __exit__(self, exc_type, exc_val, exc_tb):
        """Context manager exit."""
        if self.workbook:
            self.workbook.close()
    
    def get_sheet_names(self) -> List[str]:
        """Get list of sheet names."""
        return self.workbook.sheetnames if self.workbook else []
    
    def iter_rows(self, sheet_name: str, start_row: int = 2, max_rows: Optional[int] = None) -> Iterator[List[Any]]:
        """
        Iterate over rows in a sheet.
        
        Args:
            sheet_name: Name of sheet
            start_row: First data row (1-indexed, default 2 to skip header)
            max_rows: Maximum rows to read (None = all)
        
        Yields:
            List of cell values for each row
        """
        if not self.workbook:
            raise RuntimeError("Workbook not opened. Use as context manager.")
        
        sheet = self.workbook[sheet_name]
        end_row = sheet.max_row
        if max_rows:
            end_row = min(start_row + max_rows - 1, end_row)
        
        for row in sheet.iter_rows(min_row=start_row, max_row=end_row, values_only=True):
            yield row
    
    def get_header(self, sheet_name: str) -> List[str]:
        """
        Get header row (first row) from sheet.
        
        Args:
            sheet_name: Name of sheet
        
        Returns:
            List of column names
        """
        if not self.workbook:
            raise RuntimeError("Workbook not opened. Use as context manager.")
        
        sheet = self.workbook[sheet_name]
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        return [str(cell).strip() if cell else f"col_{i+1}" for i, cell in enumerate(header_row)]
    
    def get_row_count(self, sheet_name: str) -> int:
        """
        Get approximate row count (excluding header).
        
        Args:
            sheet_name: Name of sheet
        
        Returns:
            Row count
        """
        if not self.workbook:
            raise RuntimeError("Workbook not opened. Use as context manager.")
        
        sheet = self.workbook[sheet_name]
        return max(0, sheet.max_row - 1)  # Exclude header


def normalize_column_name(col_name: str) -> str:
    """
    Normalize column name to snake_case.
    
    Args:
        col_name: Original column name
    
    Returns:
        Normalized column name
    """
    normalized = str(col_name).strip().lower()
    normalized = normalized.replace(" ", "_")
    normalized = normalized.replace("-", "_")
    normalized = "".join(c if c.isalnum() or c == "_" else "" for c in normalized)
    # Remove leading/trailing underscores
    normalized = normalized.strip("_")
    return normalized

