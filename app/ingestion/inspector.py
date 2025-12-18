"""
Excel Inspector - Gera reports baseados nos headers REAIS do Excel.
Não inventa nada, apenas lê e analisa.
"""
import openpyxl
from pathlib import Path
from typing import Dict, Any, List, Optional, Set
from datetime import datetime
from collections import Counter
import json
import hashlib
import logging

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


class ExcelInspector:
    """Inspector que analisa o Excel real e gera reports."""
    
    def __init__(self, excel_path: str):
        """
        Initialize inspector.
        
        Args:
            excel_path: Path to Folha_IA.xlsx
        """
        self.excel_path = Path(excel_path)
        if not self.excel_path.exists():
            raise FileNotFoundError(f"Excel file not found: {excel_path}")
        
        self.workbook = None
        self.sheet_data = {}
    
    def __enter__(self):
        """Context manager entry."""
        self.workbook = openpyxl.load_workbook(
            self.excel_path,
            read_only=True,
            data_only=False
        )
        return self
    
    def __exit__(self, exc_type, exc_val, exc_tb):
        """Context manager exit."""
        if self.workbook:
            self.workbook.close()
    
    def inspect_sheet(self, sheet_name: str, max_sample_rows: int = 10000) -> Dict[str, Any]:
        """
        Inspect a single sheet.
        
        Args:
            sheet_name: Sheet name
            max_sample_rows: Max rows to sample for analysis
        
        Returns:
            Sheet inspection data
        """
        sheet = self.workbook[sheet_name]
        
        # Read header (row 1)
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        headers = [str(cell).strip() if cell else f"col_{i+1}" 
                  for i, cell in enumerate(header_row)]
        
        # Count rows (excluding header)
        total_rows = sheet.max_row - 1
        
        # Sample rows for analysis
        sample_size = min(total_rows, max_sample_rows)
        
        # Analyze columns
        column_data = {header: [] for header in headers}
        date_cells_invalid = []
        
        for row_idx, row in enumerate(sheet.iter_rows(
            min_row=2, 
            max_row=min(sheet.max_row, sample_size + 1),
            values_only=True
        ), start=2):
            for col_idx, cell_value in enumerate(row):
                if col_idx < len(headers):
                    column_data[headers[col_idx]].append(cell_value)
                    
                    # Check for invalid date cells
                    if headers[col_idx].lower() in ['data', 'inicio', 'fim', 'criacao', 'acabamento', 'transporte', 'prevista']:
                        if cell_value and not isinstance(cell_value, (datetime, type(None))):
                            try:
                                # Try to parse as date string
                                if isinstance(cell_value, str) and len(cell_value) > 0:
                                    # If it's not a valid date format, mark as invalid
                                    if not any(char.isdigit() for char in cell_value):
                                        date_cells_invalid.append({
                                            "sheet": sheet_name,
                                            "row": row_idx,
                                            "column": headers[col_idx],
                                            "value": str(cell_value)[:50]
                                        })
                            except:
                                pass
        
        # Analyze each column
        column_analysis = {}
        for header, values in column_data.items():
            non_null = [v for v in values if v is not None and str(v).strip() != '']
            null_count = len(values) - len(non_null)
            null_rate = null_count / len(values) if values else 0.0
            
            # Infer type
            inferred_type = "string"
            numeric_count = 0
            date_count = 0
            int_count = 0
            
            for val in non_null[:1000]:  # Sample
                if isinstance(val, datetime):
                    date_count += 1
                elif isinstance(val, (int, float)):
                    numeric_count += 1
                    if isinstance(val, int) or (isinstance(val, float) and val.is_integer()):
                        int_count += 1
            
            if date_count > len(non_null) * 0.8:
                inferred_type = "date"
            elif numeric_count > len(non_null) * 0.8:
                if int_count > numeric_count * 0.9:
                    inferred_type = "integer"
                else:
                    inferred_type = "float"
            
            # Cardinality
            unique_vals = set(str(v).strip() for v in non_null)
            cardinality = len(unique_vals)
            cardinality_rate = cardinality / len(non_null) if non_null else 0.0
            
            # Examples (up to 5 unique)
            examples = list(unique_vals)[:5]
            
            # Min/max for dates
            min_date = None
            max_date = None
            if inferred_type == "date":
                dates = [v for v in non_null if isinstance(v, datetime)]
                if dates:
                    min_date = min(dates).isoformat()
                    max_date = max(dates).isoformat()
            
            # Top values for categorical
            top_values = []
            if cardinality_rate < 0.1 and len(non_null) > 0:  # Low cardinality = categorical
                value_counts = Counter(str(v).strip() for v in non_null)
                top_values = [{"value": str(k), "count": v} 
                            for k, v in value_counts.most_common(10)]
            
            column_analysis[header] = {
                "inferred_type": inferred_type,
                "null_rate": round(null_rate, 4),
                "null_count": null_count,
                "cardinality": cardinality,
                "cardinality_rate": round(cardinality_rate, 4),
                "examples": examples,
                "min_date": min_date,
                "max_date": max_date,
                "top_values": top_values
            }
        
        # PK candidates (low null rate + high uniqueness)
        pk_candidates = []
        for header, analysis in column_analysis.items():
            if analysis["null_rate"] < 0.01 and analysis["cardinality_rate"] > 0.95:
                pk_candidates.append(header)
        
        return {
            "sheet_name": sheet_name,
            "headers": headers,
            "row_count": total_rows,
            "column_count": len(headers),
            "columns": column_analysis,
            "pk_candidates": pk_candidates,
            "date_cells_invalid": date_cells_invalid[:100]  # Limit to 100
        }
    
    def validate_relationships(self) -> Dict[str, Any]:
        """
        Validate relationships between sheets by checking FK matching rates.
        
        Returns:
            Relationship validation report
        """
        relationships = {
            "Of_Id ↔ FaseOf_OfId": {
                "from_sheet": "OrdensFabrico",
                "from_col": "Of_Id",
                "to_sheet": "FasesOrdemFabrico",
                "to_col": "FaseOf_OfId",
                "match_rate": None,
                "orphans": []
            },
            "Fase_Id ↔ FaseOf_FaseId": {
                "from_sheet": "Fases",
                "from_col": "Fase_Id",
                "to_sheet": "FasesOrdemFabrico",
                "to_col": "FaseOf_FaseId",
                "match_rate": None,
                "orphans": []
            },
            "Produto_Id ↔ Of_ProdutoId": {
                "from_sheet": "Modelos",
                "from_col": "Produto_Id",
                "to_sheet": "OrdensFabrico",
                "to_col": "Of_ProdutoId",
                "match_rate": None,
                "orphans": []
            },
            "Produto_Id ↔ ProdutoFase_ProdutoId": {
                "from_sheet": "Modelos",
                "from_col": "Produto_Id",
                "to_sheet": "FasesStandardModelos",
                "to_col": "ProdutoFase_ProdutoId",
                "match_rate": None,
                "orphans": []
            },
            "Fase_Id ↔ ProdutoFase_FaseId": {
                "from_sheet": "Fases",
                "from_col": "Fase_Id",
                "to_sheet": "FasesStandardModelos",
                "to_col": "ProdutoFase_FaseId",
                "match_rate": None,
                "orphans": []
            },
            "Funcionario_Id ↔ FuncionarioFase_FuncionarioId": {
                "from_sheet": "Funcionarios",
                "from_col": "Funcionario_Id",
                "to_sheet": "FuncionariosFasesAptos",
                "to_col": "FuncionarioFase_FuncionarioId",
                "match_rate": None,
                "orphans": []
            },
            "Funcionario_Id ↔ FuncionarioFaseOf_FuncionarioId": {
                "from_sheet": "Funcionarios",
                "from_col": "Funcionario_Id",
                "to_sheet": "FuncionariosFaseOrdemFabrico",
                "to_col": "FuncionarioFaseOf_FuncionarioId",
                "match_rate": None,
                "orphans": []
            },
            "FuncionarioFaseOf_FaseOfId ↔ FaseOf_Id": {
                "from_sheet": "FuncionariosFaseOrdemFabrico",
                "from_col": "FuncionarioFaseOf_FaseOfId",
                "to_sheet": "FasesOrdemFabrico",
                "to_col": "FaseOf_Id",
                "match_rate": None,
                "orphans": [],
                "note": "CRITICAL: Validating if FuncionarioFaseOf_FaseOfId maps to FaseOf_Id"
            }
        }
        
        # Collect all sheet data
        sheet_values = {}
        for sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
            header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
            headers = [str(cell).strip() if cell else f"col_{i+1}" 
                      for i, cell in enumerate(header_row)]
            
            # Read all values for each column
            column_values = {header: set() for header in headers}
            for row in sheet.iter_rows(min_row=2, max_row=min(sheet.max_row, 100000), values_only=True):
                for col_idx, cell_value in enumerate(row):
                    if col_idx < len(headers) and cell_value is not None:
                        column_values[headers[col_idx]].add(str(cell_value).strip())
            
            sheet_values[sheet_name] = column_values
        
        # Validate each relationship
        for rel_name, rel_data in relationships.items():
            from_sheet = rel_data["from_sheet"]
            from_col = rel_data["from_col"]
            to_sheet = rel_data["to_sheet"]
            to_col = rel_data["to_col"]
            
            if from_sheet not in sheet_values or to_sheet not in sheet_values:
                rel_data["match_rate"] = None
                rel_data["error"] = f"Sheet not found: {from_sheet} or {to_sheet}"
                continue
            
            if from_col not in sheet_values[from_sheet] or to_col not in sheet_values[to_sheet]:
                rel_data["match_rate"] = None
                rel_data["error"] = f"Column not found: {from_col} or {to_col}"
                continue
            
            from_values = sheet_values[from_sheet][from_col]
            to_values = sheet_values[to_sheet][to_col]
            
            # Calculate match rate
            matches = from_values.intersection(to_values)
            total_to = len(to_values)
            match_rate = len(matches) / total_to if total_to > 0 else 0.0
            
            rel_data["match_rate"] = round(match_rate, 4)
            rel_data["matches"] = len(matches)
            rel_data["total_from"] = len(from_values)
            rel_data["total_to"] = total_to
            
            # Find orphans (values in 'to' that don't exist in 'from')
            orphans = to_values - from_values
            rel_data["orphans"] = list(orphans)[:100]  # Limit to 100
            rel_data["orphan_count"] = len(orphans)
        
        return relationships
    
    def generate_reports(self, output_dir: Path):
        """
        Generate all inspection reports.
        
        Args:
            output_dir: Output directory for reports
        """
        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)
        
        logger.info(f"Starting inspection of {self.excel_path}")
        
        # Inspect all sheets
        sheet_inspections = {}
        for sheet_name in self.workbook.sheetnames:
            logger.info(f"Inspecting sheet: {sheet_name}")
            sheet_inspections[sheet_name] = self.inspect_sheet(sheet_name)
        
        # Validate relationships
        logger.info("Validating relationships")
        relationships = self.validate_relationships()
        
        # Generate DATA_DICTIONARY.md
        self._generate_data_dictionary(output_dir, sheet_inspections)
        
        # Generate PROFILE_REPORT.json
        self._generate_profile_report(output_dir, sheet_inspections)
        
        # Generate RELATIONSHIPS_REPORT.json
        self._generate_relationships_report(output_dir, relationships)
        
        logger.info(f"Inspection complete. Reports in {output_dir}")
    
    def _generate_data_dictionary(self, output_dir: Path, sheet_inspections: Dict[str, Any]):
        """Generate DATA_DICTIONARY.md"""
        md_lines = [
            "# Data Dictionary",
            "",
            f"**Generated at:** {datetime.now().isoformat()}",
            f"**Source file:** {self.excel_path}",
            "",
            "---",
            ""
        ]
        
        for sheet_name, inspection in sheet_inspections.items():
            md_lines.extend([
                f"## Sheet: `{sheet_name}`",
                "",
                f"- **Row count:** {inspection['row_count']:,}",
                f"- **Column count:** {inspection['column_count']}",
                ""
            ])
            
            if inspection.get("pk_candidates"):
                md_lines.append(f"- **Primary key candidates:** {', '.join(inspection['pk_candidates'])}")
                md_lines.append("")
            
            md_lines.append("### Columns")
            md_lines.append("")
            md_lines.append("| Column Name | Type | Null Rate | Null Count | Cardinality | Cardinality Rate | Examples | Min Date | Max Date |")
            md_lines.append("|-------------|------|-----------|------------|-------------|------------------|----------|----------|----------|")
            
            for col_name, col_data in inspection["columns"].items():
                examples_str = ", ".join(str(ex)[:30] for ex in col_data["examples"][:3])
                if len(examples_str) > 50:
                    examples_str = examples_str[:47] + "..."
                
                min_date = col_data.get("min_date", "")
                max_date = col_data.get("max_date", "")
                
                md_lines.append(
                    f"| `{col_name}` | {col_data['inferred_type']} | "
                    f"{col_data['null_rate']:.2%} | {col_data['null_count']:,} | "
                    f"{col_data['cardinality']:,} | {col_data['cardinality_rate']:.2%} | "
                    f"{examples_str} | {min_date} | {max_date} |"
                )
            
            md_lines.append("")
            md_lines.append("---")
            md_lines.append("")
        
        output_path = output_dir / "DATA_DICTIONARY.md"
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write('\n'.join(md_lines))
        
        logger.info(f"Data dictionary generated: {output_path}")
    
    def _generate_profile_report(self, output_dir: Path, sheet_inspections: Dict[str, Any]):
        """Generate PROFILE_REPORT.json"""
        profile = {
            "generated_at": datetime.now().isoformat(),
            "excel_path": str(self.excel_path),
            "excel_sha256": self._calculate_file_hash(),
            "sheets": {}
        }
        
        for sheet_name, inspection in sheet_inspections.items():
            profile["sheets"][sheet_name] = {
                "row_count_real": inspection["row_count"],
                "column_count": inspection["column_count"],
                "headers": inspection["headers"],
                "columns": {
                    col_name: {
                        "null_rate": col_data["null_rate"],
                        "null_count": col_data["null_count"],
                        "distinct_approx": col_data["cardinality"],
                        "inferred_type": col_data["inferred_type"],
                        "min_date": col_data.get("min_date"),
                        "max_date": col_data.get("max_date"),
                        "top_values": col_data.get("top_values", [])
                    }
                    for col_name, col_data in inspection["columns"].items()
                },
                "date_cells_invalid": inspection.get("date_cells_invalid", []),
                "pk_candidates": inspection.get("pk_candidates", [])
            }
        
        output_path = output_dir / "PROFILE_REPORT.json"
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(profile, f, indent=2, default=str)
        
        logger.info(f"Profile report generated: {output_path}")
    
    def _generate_relationships_report(self, output_dir: Path, relationships: Dict[str, Any]):
        """Generate RELATIONSHIPS_REPORT.json"""
        report = {
            "generated_at": datetime.now().isoformat(),
            "relationships": relationships,
            "summary": {
                "total_relationships": len(relationships),
                "validated": len([r for r in relationships.values() if r.get("match_rate") is not None]),
                "high_match_rate": len([r for r in relationships.values() 
                                       if r.get("match_rate") and r["match_rate"] > 0.99]),
                "low_match_rate": len([r for r in relationships.values() 
                                      if r.get("match_rate") and r["match_rate"] < 0.9])
            }
        }
        
        output_path = output_dir / "RELATIONSHIPS_REPORT.json"
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(report, f, indent=2, default=str)
        
        logger.info(f"Relationships report generated: {output_path}")
    
    def _calculate_file_hash(self) -> str:
        """Calculate SHA256 hash of Excel file."""
        sha256 = hashlib.sha256()
        with open(self.excel_path, 'rb') as f:
            for chunk in iter(lambda: f.read(4096), b""):
                sha256.update(chunk)
        return sha256.hexdigest()


def main():
    """Main entry point."""
    import sys
    from pathlib import Path
    
    excel_path = Path(__file__).parent.parent.parent / "data" / "raw" / "Folha_IA.xlsx"
    output_dir = Path(__file__).parent
    
    if not excel_path.exists():
        print(f"Error: Excel file not found at {excel_path}")
        sys.exit(1)
    
    with ExcelInspector(str(excel_path)) as inspector:
        inspector.generate_reports(output_dir)
    
    print(f"\n✅ Inspection complete! Reports generated in {output_dir}")
    print(f"  - DATA_DICTIONARY.md")
    print(f"  - PROFILE_REPORT.json")
    print(f"  - RELATIONSHIPS_REPORT.json")


if __name__ == "__main__":
    main()

