"""
Extract Phase: Convert Excel to CSV.gz files.
Memory-efficient, streaming approach.
"""
import openpyxl
import csv
import gzip
from pathlib import Path
from typing import Dict, Any, Optional
import hashlib
import structlog
from datetime import datetime

logger = structlog.get_logger()


class ExcelExtractor:
    """Extract Excel sheets to CSV.gz files."""
    
    def __init__(self, excel_path: str, output_dir: Path):
        """
        Initialize extractor.
        
        Args:
            excel_path: Path to Excel file
            output_dir: Output directory for CSV.gz files
        """
        self.excel_path = Path(excel_path)
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.workbook = None
        self.sheet_checksums = {}
    
    def __enter__(self):
        """Context manager entry."""
        self.workbook = openpyxl.load_workbook(
            self.excel_path,
            read_only=True,
            data_only=False
        )
        return self
    
    def __exit__(self, exc_type, exc_val, exc_tb):
        """Context manager exit."""
        if self.workbook:
            self.workbook.close()
    
    def extract_sheet(self, sheet_name: str) -> Dict[str, Any]:
        """
        Extract a sheet to CSV.gz.
        
        Args:
            sheet_name: Name of sheet to extract
        
        Returns:
            Dict with file path, row count, checksum
        """
        logger.info(f"Extracting sheet: {sheet_name}")
        
        sheet = self.workbook[sheet_name]
        
        # Read header
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        headers = [str(cell).strip() if cell else f"col_{i+1}" 
                  for i, cell in enumerate(header_row)]
        
        # Output file
        csv_gz_path = self.output_dir / f"{sheet_name}.csv.gz"
        
        # Extract to CSV.gz
        row_count = 0
        sha256 = hashlib.sha256()
        
        with gzip.open(csv_gz_path, 'wt', encoding='utf-8', compresslevel=6) as gz_file:
            writer = csv.writer(gz_file, quoting=csv.QUOTE_MINIMAL)
            
            # Write header
            writer.writerow(headers)
            sha256.update(','.join(headers).encode('utf-8'))
            
            # Write rows
            for row in sheet.iter_rows(min_row=2, values_only=True):
                # Normalize values
                normalized_row = []
                for cell_value in row:
                    if cell_value is None:
                        normalized_row.append('')
                    elif isinstance(cell_value, datetime):
                        normalized_row.append(cell_value.isoformat())
                    else:
                        normalized_row.append(str(cell_value))
                
                writer.writerow(normalized_row)
                row_count += 1
                
                # Update checksum
                row_str = ','.join(normalized_row)
                sha256.update(row_str.encode('utf-8'))
        
        checksum = sha256.hexdigest()
        self.sheet_checksums[sheet_name] = checksum
        
        file_size_mb = csv_gz_path.stat().st_size / (1024 * 1024)
        
        logger.info(
            f"Extracted {sheet_name}",
            row_count=row_count,
            file_size_mb=round(file_size_mb, 2),
            checksum=checksum
        )
        
        return {
            'sheet_name': sheet_name,
            'file_path': str(csv_gz_path),
            'row_count': row_count,
            'checksum': checksum,
            'file_size_mb': round(file_size_mb, 2)
        }
    
    def extract_all(self) -> Dict[str, Any]:
        """
        Extract all sheets.
        
        Returns:
            Dict with all extraction results and overall checksum
        """
        results = {}
        total_rows = 0
        
        for sheet_name in self.workbook.sheetnames:
            sheet_result = self.extract_sheet(sheet_name)
            results[sheet_name] = sheet_result
            total_rows += sheet_result.get('row_count', 0)
        
        # Calculate overall Excel checksum
        excel_checksum = self._calculate_excel_checksum()
        
        # Calculate per-sheet checksums summary
        per_sheet_checksums = {name: data['checksum'] for name, data in results.items()}
        
        return {
            'excel_path': str(self.excel_path),
            'excel_checksum': excel_checksum,
            'per_sheet_sha256': per_sheet_checksums,
            'sheets': results,
            'total_rows_extracted': total_rows,
            'extracted_at': datetime.now().isoformat()
        }
    
    def _calculate_excel_checksum(self) -> str:
        """Calculate SHA256 of Excel file."""
        sha256 = hashlib.sha256()
        with open(self.excel_path, 'rb') as f:
            for chunk in iter(lambda: f.read(4096), b""):
                sha256.update(chunk)
        return sha256.hexdigest()


def main():
    """CLI entry point."""
    import sys
    from pathlib import Path
    
    excel_path = Path(__file__).parent.parent.parent / "data" / "raw" / "Folha_IA.xlsx"
    output_dir = Path(__file__).parent.parent.parent / "data" / "processed"
    
    if not excel_path.exists():
        print(f"Error: Excel file not found at {excel_path}")
        sys.exit(1)
    
    structlog.configure(
        processors=[
            structlog.processors.TimeStamper(fmt="iso"),
            structlog.processors.JSONRenderer()
        ]
    )
    
    with ExcelExtractor(str(excel_path), output_dir) as extractor:
        results = extractor.extract_all()
        
        # Save extraction report
        import json
        report_path = output_dir / "extraction_report.json"
        with open(report_path, 'w') as f:
            json.dump(results, f, indent=2)
        
        print(f"\n✅ Extraction complete!")
        print(f"  Excel checksum: {results['excel_checksum']}")
        print(f"  Sheets extracted: {len(results['sheets'])}")
        print(f"  Report saved: {report_path}")


if __name__ == "__main__":
    main()

