"""
Inspect Excel file and generate DATA_DICTIONARY.md with real schema analysis.
Uses openpyxl in read_only mode for memory efficiency.
"""
import openpyxl
from pathlib import Path
from typing import Dict, List, Any, Optional
from collections import Counter
from datetime import datetime
import json
import re


def infer_type_from_values(values: List[Any]) -> Dict[str, Any]:
    """Infer data type, null rate, and examples from a column's values."""
    non_null = [v for v in values if v is not None and str(v).strip() != '']
    null_count = len(values) - len(non_null)
    null_rate = null_count / len(values) if values else 0.0
    
    if not non_null:
        return {
            "inferred_type": "unknown",
            "null_rate": null_rate,
            "null_count": null_count,
            "cardinality": 0,
            "examples": []
        }
    
    # Try to infer type
    inferred_type = "string"
    numeric_count = 0
    date_count = 0
    int_count = 0
    
    for val in non_null[:1000]:  # Sample first 1000
        val_str = str(val).strip()
        if val_str == '':
            continue
        
        # Check if numeric
        try:
            float(val_str.replace(',', '').replace(' ', ''))
            numeric_count += 1
            if '.' not in val_str and 'e' not in val_str.lower():
                int_count += 1
        except:
            pass
        
        # Check if date-like
        if isinstance(val, datetime):
            date_count += 1
        elif re.match(r'\d{4}[-/]\d{1,2}[-/]\d{1,2}', val_str):
            date_count += 1
    
    if date_count > len(non_null) * 0.8:
        inferred_type = "date"
    elif numeric_count > len(non_null) * 0.8:
        if int_count > numeric_count * 0.9:
            inferred_type = "integer"
        else:
            inferred_type = "float"
    else:
        inferred_type = "string"
    
    # Calculate cardinality (unique values)
    unique_vals = set(str(v).strip() for v in non_null)
    cardinality = len(unique_vals)
    
    # Get examples (up to 5 unique)
    examples = list(unique_vals)[:5]
    
    return {
        "inferred_type": inferred_type,
        "null_rate": round(null_rate, 4),
        "null_count": null_count,
        "cardinality": cardinality,
        "cardinality_rate": round(cardinality / len(non_null), 4) if non_null else 0.0,
        "examples": examples
    }


def analyze_sheet(ws: openpyxl.worksheet.worksheet.Worksheet, max_rows: int = 10000) -> Dict[str, Any]:
    """Analyze a single sheet and return column metadata."""
    # Read header
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    if not header_row or not any(header_row):
        return {}
    
    columns = [str(cell).strip() if cell else f"col_{i+1}" for i, cell in enumerate(header_row)]
    
    # Read sample rows (up to max_rows for analysis)
    rows_to_analyze = min(ws.max_row - 1, max_rows)
    
    column_data: Dict[str, List[Any]] = {col: [] for col in columns}
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=min(ws.max_row, max_rows + 1), values_only=True), 1):
        for col_idx, cell_value in enumerate(row):
            if col_idx < len(columns):
                column_data[columns[col_idx]].append(cell_value)
    
    # Analyze each column
    column_analysis = {}
    for col_name, values in column_data.items():
        analysis = infer_type_from_values(values)
        column_analysis[col_name] = analysis
    
    # Infer potential primary keys (low null rate + high uniqueness)
    pk_candidates = []
    for col_name, analysis in column_analysis.items():
        if analysis["null_rate"] < 0.01 and analysis["cardinality_rate"] > 0.95:
            pk_candidates.append(col_name)
    
    return {
        "row_count": ws.max_row - 1,  # Exclude header
        "column_count": len(columns),
        "columns": column_analysis,
        "pk_candidates": pk_candidates
    }


def inspect_excel(file_path: str) -> Dict[str, Any]:
    """Inspect entire Excel file and return comprehensive analysis."""
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=False)
    
    analysis = {
        "file_path": file_path,
        "inspected_at": datetime.now().isoformat(),
        "sheets": {}
    }
    
    for sheet_name in wb.sheetnames:
        print(f"Analyzing sheet: {sheet_name}...")
        ws = wb[sheet_name]
        sheet_analysis = analyze_sheet(ws)
        analysis["sheets"][sheet_name] = sheet_analysis
    
    wb.close()
    return analysis


def generate_data_dictionary_markdown(analysis: Dict[str, Any], output_path: str):
    """Generate DATA_DICTIONARY.md from analysis."""
    md_lines = [
        "# Data Dictionary",
        "",
        f"**Generated at:** {analysis['inspected_at']}",
        f"**Source file:** {analysis['file_path']}",
        "",
        "---",
        ""
    ]
    
    for sheet_name, sheet_data in analysis["sheets"].items():
        md_lines.extend([
            f"## Sheet: `{sheet_name}`",
            "",
            f"- **Row count:** {sheet_data['row_count']:,}",
            f"- **Column count:** {sheet_data['column_count']}",
            ""
        ])
        
        if sheet_data.get("pk_candidates"):
            md_lines.append(f"- **Primary key candidates:** {', '.join(sheet_data['pk_candidates'])}")
            md_lines.append("")
        
        md_lines.append("### Columns")
        md_lines.append("")
        md_lines.append("| Column Name | Type | Null Rate | Null Count | Cardinality | Cardinality Rate | Examples |")
        md_lines.append("|-------------|------|-----------|------------|-------------|------------------|----------|")
        
        for col_name, col_data in sheet_data["columns"].items():
            examples_str = ", ".join(str(ex)[:30] for ex in col_data["examples"][:3])
            if len(examples_str) > 50:
                examples_str = examples_str[:47] + "..."
            
            md_lines.append(
                f"| `{col_name}` | {col_data['inferred_type']} | "
                f"{col_data['null_rate']:.2%} | {col_data['null_count']:,} | "
                f"{col_data['cardinality']:,} | {col_data['cardinality_rate']:.2%} | "
                f"{examples_str} |"
            )
        
        md_lines.append("")
        md_lines.append("---")
        md_lines.append("")
    
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write('\n'.join(md_lines))
    
    print(f"Data dictionary written to: {output_path}")


def main():
    """Main entry point."""
    excel_path = Path(__file__).parent.parent.parent / "data" / "raw" / "Folha_IA.xlsx"
    
    if not excel_path.exists():
        print(f"Error: Excel file not found at {excel_path}")
        return
    
    print(f"Inspecting {excel_path}...")
    analysis = inspect_excel(str(excel_path))
    
    # Save JSON analysis
    json_path = Path(__file__).parent.parent.parent / "app" / "ingestion" / "excel_analysis.json"
    json_path.parent.mkdir(parents=True, exist_ok=True)
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(analysis, f, indent=2, default=str)
    print(f"JSON analysis saved to: {json_path}")
    
    # Generate markdown
    md_path = Path(__file__).parent.parent.parent / "app" / "ingestion" / "DATA_DICTIONARY.md"
    generate_data_dictionary_markdown(analysis, str(md_path))


if __name__ == "__main__":
    main()

